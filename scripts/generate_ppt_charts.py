from pathlib import Path
import pandas as pd
import matplotlib.pyplot as plt
from openpyxl import load_workbook


ROOT_DIR = Path(__file__).resolve().parent.parent
DATA_DIR = ROOT_DIR / "data"
OUTPUT_DIR = ROOT_DIR / "assets" / "figures"
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)


def read_excel_basic(path: Path) -> pd.DataFrame:
    workbook = load_workbook(path, data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    headers = rows[0]
    data = rows[1:]
    return pd.DataFrame(data, columns=headers)


def plot_slide_7_job_postings() -> None:
    df = read_excel_basic(DATA_DIR / "job-postings.xlsx")
    df = df.sort_values("Number of Job Postings", ascending=False)

    plt.figure(figsize=(10, 6))
    colors = plt.cm.Blues([0.35 + i * 0.08 for i in range(len(df))])
    plt.bar(df["Location"], df["Number of Job Postings"], color=colors)
    plt.title("Job Postings by Location")
    plt.xlabel("Location")
    plt.ylabel("Number of Job Postings")
    plt.xticks(rotation=30, ha="right")

    for index, value in enumerate(df["Number of Job Postings"]):
        plt.text(index, value + max(df["Number of Job Postings"]) * 0.01, f"{value:,}", ha="center", va="bottom", fontsize=9)

    plt.tight_layout()
    plt.savefig(OUTPUT_DIR / "slide_7_job_postings.png", dpi=300, bbox_inches="tight")
    plt.close()


def plot_slide_8_popular_languages() -> None:
    df = pd.read_csv(DATA_DIR / "popular-languages.csv")
    df["SalaryNumeric"] = df["Average Annual Salary"].replace({"[$,]": ""}, regex=True).astype(int)
    df = df.sort_values("SalaryNumeric", ascending=False)

    plt.figure(figsize=(10, 6))
    colors = plt.cm.Greens([0.35 + i * 0.05 for i in range(len(df))])
    plt.bar(df["Language"], df["SalaryNumeric"], color=colors)
    plt.title("Average Annual Salary by Programming Language")
    plt.xlabel("Language")
    plt.ylabel("Average Annual Salary (USD)")
    plt.xticks(rotation=30, ha="right")

    for index, value in enumerate(df["SalaryNumeric"]):
        plt.text(index, value + max(df["SalaryNumeric"]) * 0.01, f"${value:,.0f}", ha="center", va="bottom", fontsize=9)

    plt.tight_layout()
    plt.savefig(OUTPUT_DIR / "slide_8_popular_languages_salary.png", dpi=300, bbox_inches="tight")
    plt.close()


def main() -> None:
    plt.style.use("ggplot")
    plot_slide_7_job_postings()
    plot_slide_8_popular_languages()
    print(f"Charts saved to: {OUTPUT_DIR}")
    print("- slide_7_job_postings.png")
    print("- slide_8_popular_languages_salary.png")


if __name__ == "__main__":
    main()
